import sys
import re
import math
from openpyxl import load_workbook, styles


def make_tabs(file_name):
    converted = load_workbook(file_name)
    converted.create_sheet('Access Required')
    converted.create_sheet('Infrastructure Required')
    converted.create_sheet('Tool Required')
    converted.create_sheet('Simple Workload')
    converted.create_sheet('Data')
    converted.save(file_name)
    return converted
def make_copy(file_name, filename2, sheet_name, sheet, name, keywords):
    fle = load_workbook(file_name)
    fill_pattern = styles.PatternFill(patternType='solid', fgColor='DDEBF7')
    ar1 = fle[sheet_name]
    mr = ar1.max_row
    mc = ar1.max_column
    for i in range(1,mr+1):
        for j in range(1,mc+1):
            c = ar1.cell(row=i, column=j)
            sheet[name].cell(row=i, column=j).value = c.value
            if c.value in keywords:
                sheet[name].cell(row=i, column=j).font = styles.Font(bold=True)
                sheet[name].cell(row=i, column=j).fill = fill_pattern
    dims={}
    mr1 = sheet[name].max_row
    mc1 = sheet[name].max_column
    for i in range(1,mc1+1):
        chk = sheet[name].cell(row=1, column=i)
        chk2 = sheet[name].cell(row=2, column=i)
        if chk.value:
            chk.font = styles.Font(bold=True)
            chk.fill = fill_pattern
        if chk2.value:
            chk2.font = styles.Font(bold=True)
            chk2.fill = fill_pattern
        for j in range(1,mr1+1):

            if chk.value or chk2.value:
                all_border = styles.Border(left=styles.Side(style='thin'), 
                right=styles.Side(style='thin'), 
                top=styles.Side(style='thin'), 
                bottom=styles.Side(style='thin'))
                sheet[name].cell(row=j, column=i).border = all_border
    for row in sheet[name].rows:
        for cell in row:
            
            
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
   
    for col, value in dims.items():
        sheet[name].column_dimensions[col].width = value
    sheet.save(filename2)
    return sheet
def data_transfer(sheet, name, name2, keyword, filename, rge):
    m_r = sheet[name].max_row
    
    m_c = sheet[name].max_column
   
    for k in range(1, m_r+1):
        val = sheet[name].cell(row=k,column=1).value
        if  val == keyword:
            m_t = sheet[name].cell(row=k,column=9).value
            for l in range(rge[0],rge[1]):
                if m_t == sheet[name2].cell(row=l, column=1).value:
                    ar_l = sheet[name].cell(row=k, column=20).value
                    if ar_l == "I":
                        ar_val = sheet[name2].cell(row=l, column=2)
                        if ar_val.value:
                            ar_val.value += 1
                            
                        else:
                            ar_val.value = 1
                    elif ar_l == "I,R":
                        ar_val = sheet[name2].cell(row=l, column=3)
                        if ar_val.value:
                            ar_val.value += 1
                        else:
                            ar_val.value = 1

                    elif ar_l == "I,S":
                        ar_val = sheet[name2].cell(row=l, column=4)
                        if ar_val.value:
                            ar_val.value += 1
                        else:
                            ar_val.value = 1
                    elif ar_l == "U":
                        ar_val = sheet[name2].cell(row=l, column=5)
                        if ar_val.value:
                            ar_val.value += 1
                            
                        else:
                            ar_val.value = 1
                    elif ar_l == "U,I":
                        ar_val = sheet[name2].cell(row=l, column=6)
                        if ar_val.value:
                            ar_val.value += 1
                        else:
                            ar_val.value = 1
                    elif ar_l == "U,I,R":
                        ar_val = sheet[name2].cell(row=l, column=7)
                        
                        if ar_val.value:
                            ar_val.value += 1
                            
                        else:
                            ar_val.value = 1
                    elif ar_l == "U,S":
                        ar_val = sheet[name2].cell(row=l, column=8)
                        if ar_val.value:
                            ar_val.value += 1
                        else:
                            ar_val.value = 1
                    elif ar_l == "U,S,R":
                        ar_val = sheet[name2].cell(row=l, column=9)
                        if ar_val.value:
                            ar_val.value += 1
                        else:
                            ar_val.value = 1
                    elif ar_l == "S":
                        ar_val = sheet[name2].cell(row=l, column=10)
                        if ar_val.value:
                            ar_val.value += 1
                        else:
                            ar_val.value = 1
    sheet.save(filename)
    return sheet


def inf_transfer(sheet, name, name2, keyword, filename, rge):
    m_r = sheet[name].max_row
    m_c = sheet[name].max_column
    for k in range(1, m_r+1):
        val = sheet[name].cell(row=k,column=1).value
        if  val == keyword:
            m_t = sheet[name].cell(row=k,column=9).value
            
            for l in range(rge[0],rge[1]):
                if m_t == sheet[name2].cell(row=l, column=1).value:
                    ar_1 = sheet[name].cell(row=k, column=21).value
                    ar_2 = sheet[name].cell(row=k, column=23).value
                    ar_3 = sheet[name].cell(row=k, column=25).value
                    ct = 0
                    while ct != 3:
                        ct += 1
                        if ct == 1:
                            if ar_1 == "E":
                                ar_v = sheet[name2].cell(row=l, column=2)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_1 == "S":
                                ar_v = sheet[name2].cell(row=l, column=4)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_1 == "P":
                                ar_v = sheet[name2].cell(row=l, column=3)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                        elif ct == 2:
                            if ar_2 == "E":
                                ar_v = sheet[name2].cell(row=l, column=8)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_2 == "S":
                                ar_v = sheet[name2].cell(row=l, column=10)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_2 == "P":
                                ar_v = sheet[name2].cell(row=l, column=9)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                        elif ct == 3:
                            if ar_3 == "E":
                                ar_v = sheet[name2].cell(row=l, column=14)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_3 == "S":
                                ar_v = sheet[name2].cell(row=l, column=16)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_3 == "P":
                                ar_v = sheet[name2].cell(row=l, column=15)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                        
    sheet.save(filename)
    return sheet


def tool_transfer(sheet, name, name2, keyword, filename, rge):
    m_r = sheet[name].max_row
    m_c = sheet[name].max_column
    for k in range(1, m_r+1):
        val = sheet[name].cell(row=k,column=1).value
        if  val == keyword:
            m_t = sheet[name].cell(row=k,column=9).value
            
            for l in range(rge[0],rge[1]):
                if m_t == sheet[name2].cell(row=l, column=1).value:
                    ar_1 = sheet[name].cell(row=k, column=22).value
                    ar_2 = sheet[name].cell(row=k, column=24).value
                    ar_3 = sheet[name].cell(row=k, column=26).value
                    ct = 0
                    while ct != 3:
                        ct += 1
                        if ct == 1:
                            if ar_1 == "C":
                                ar_v = sheet[name2].cell(row=l, column=2)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_1 == "G":
                                ar_v = sheet[name2].cell(row=l, column=3)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_1 == "L":
                                ar_v = sheet[name2].cell(row=l, column=4)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_1 == "L,C":
                                ar_v = sheet[name2].cell(row=l, column=5)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_1 == "Not Required":
                                ar_v = sheet[name2].cell(row=l, column=6)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                        elif ct == 2:
                            if ar_2 == "C":
                                ar_v = sheet[name2].cell(row=l, column=9)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_2 == "G":
                                ar_v = sheet[name2].cell(row=l, column=10)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_2 == "L":
                                ar_v = sheet[name2].cell(row=l, column=11)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_2 == "L,G":
                                ar_v = sheet[name2].cell(row=l, column=12)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_2 == "Not Required":
                                ar_v = sheet[name2].cell(row=l, column=13)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                        elif ct == 3:
                            if ar_3 == "C":
                                ar_v = sheet[name2].cell(row=l, column=17)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_3 == "G":
                                ar_v = sheet[name2].cell(row=l, column=18)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_3 == "L":
                                ar_v = sheet[name2].cell(row=l, column=19)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_2 == "L,G":
                                ar_v = sheet[name2].cell(row=l, column=20)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                            elif ar_2 == "Not Required":
                                ar_v = sheet[name2].cell(row=l, column=21)
                                if ar_v.value:
                                    ar_v.value += 1
                                else:
                                    ar_v.value = 1
                        
    sheet.save(filename)
    return sheet

def simp_wkl(sheet, name, name2, keyword, filename, rge, op_values, op_key): 
    m_r = sheet[name].max_row
    m_c = sheet[name].max_column
    op = {'/': lambda x, y: x / y,
      '*': lambda x, y: x * y}
    for j in range(rge[0],rge[1]):
        k_w = sheet[name2].cell(row=j,column=1).value
        k_count = 0
        for k in range(1, m_r+1):
            val = sheet[name].cell(row=k,column=1).value
            m_t = sheet[name].cell(row=k,column=9).value
            if  val == keyword and m_t == k_w:
                k_count +=1
        tasks = sheet[name2].cell(row=j,column=2)
        tasks.value = k_count
    ct = 0
    for n in range(rge[0],rge[1]):
        m_wkl = sheet[name2].cell(row=n,column=3)
        m_wkl.value = round(op[op_key[ct]](sheet[name2].cell(row=n,column=2).value,op_values[ct]),2)
        
        wkl_y = sheet[name2].cell(row=n,column=6)
        wkl_y.value = round(m_wkl.value/sheet[name2].cell(row=n,column=4).value,2)
        t_wkl = sheet[name2].cell(row=n,column=7)
        t_wkl.value = round(wkl_y.value * sheet[name].cell(row=2,column=27).value,2)
        ct+=1
    sheet.save(filename)
    return sheet
def optimise(elevated, pit, slab):
    t_s = round(elevated + pit + slab, 1)
    ts_d = math.modf(t_s)
    return ts_d



def data(sheet, name, name2, name3, name4, filename, name5):
    num_l = [[3,7],[9,12],[14,19],[21,24],[26,27],[29,36],[38,44]]
    dst = [2,3,4]
    r_l= [3,5,7]
    p_l = [10,12,14]
    p_2 = [21,19,17]
    alp = ["0","A","B","C","D","E","F","G","H","I", "J","K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    for i_i in num_l:
        for i in range(i_i[0], i_i[1]+1):
            impt = 0
            r_d = 0
            for j in dst:
                if sheet[name2].cell(row=i,column=j).value != None:
                    if j == 2:
                        impt = 30
                        r_d = 3
                        break
                    elif j == 3:
                        impt = 31
                        r_d = 5
                        break
                    else:
                        impt = 32
                        r_d = 7
                        break
            tools_val = 0
            val = round(sheet[name3].cell(row=i, column=7).value/sheet[name].cell(row=2,column=impt).value,1)
            for tools in range(2,7):
                if sheet[name5].cell(row=i, column=tools).value != None:
                    tools_val = sheet[name5].cell(row=1, column=tools).value + ", " + str(sheet[name5].cell(row=i, column=tools).value)
                    break
            raw_d = sheet[name4].cell(row=i,column=r_d)
            d_tool = sheet[name4].cell(row=i,column=r_d +1)
            ps_1 = sheet[name4].cell(row=i,column=10)
            ps_1_tool = sheet[name4].cell(row=i,column=11)
            ps_2 = sheet[name4].cell(row=i,column=21)
            ps_2_tool = sheet[name4].cell(row=i,column=22)
            raw_d.value = val
            d_tool.value = tools_val
            ps_1.value = val
            ps_1_tool.value = tools_val
            ps_2.value = val
            ps_2_tool.value = tools_val
    for r in r_l:
        
        net_r = sheet[name4].cell(row=46, column=r)
        net_r.value=0
        cal_r = sheet[name4].cell(row=47, column=r)
        design_r = sheet[name4].cell(row=48, column=r)
        for ad in range(3,45):
            if sheet[name4].cell(row=ad, column=r).value != None:
                net_r.value += sheet[name4].cell(row=ad, column=r).value
        net_r.value /= round(sheet[name].cell(row=2,column=28).value,1)
        net_r.value = round(net_r.value,1)
        cal_r.value = round(1.3*net_r.value,1)
        
        design_r.value = round(1.2*cal_r.value,1)
        
        
        if (design_r.value).is_integer() == False:
            lg_rd_num = int(math.ceil(design_r.value))
        sheet[name4].cell(row=49,column=r).value = lg_rd_num
    
    deduct = optimise(sheet[name4].cell(row=48,column=3).value,sheet[name4].cell(row=48,column=5).value,sheet[name4].cell(row=48,column=7).value)
    
    extras = round((deduct[1]/1.56) * round(sheet[name].cell(row=2,column=28).value,1),1)
    tt_deduct = round((deduct[0]/1.56) * round(sheet[name].cell(row=2,column=28).value,1),1)
    print(tt_deduct)
    sheet.save(filename)
    while tt_deduct > 0:
        
        lgst = 0
        smst = 100
        lg_rw = 0
        for n_l in num_l:
            for vl in range(n_l[0],n_l[1]+1):
                if sheet[name4].cell(row=vl, column=10).value:
                    if sheet[name4].cell(row=vl, column=10).value > lgst:
                        lgst = sheet[name4].cell(row=vl, column=10).value
                        lg_rw = vl
                       
                        
                    elif sheet[name4].cell(row=vl, column=10).value < smst:
                        smst = sheet[name4].cell(row=vl, column=10).value
                        sm_rw = vl
                        
        if tt_deduct - lgst < 0:
            
            sheet[name4].cell(row=sm_rw, column=14).value = smst
            sheet[name4].cell(row=sm_rw, column=15).value = sheet[name4].cell(row=sm_rw, column=11).value
            sheet[name4].cell(row=sm_rw, column=10).value = None
            sheet[name4].cell(row=sm_rw, column=11).value = None
            sheet[name4].cell(row=sm_rw, column=19).value = smst
            sheet[name4].cell(row=sm_rw, column=20).value = sheet[name4].cell(row=sm_rw, column=22).value
            sheet[name4].cell(row=sm_rw, column=21).value = None
            sheet[name4].cell(row=sm_rw, column=22).value = None
            
            tt_deduct -= smst
            sheet.save(filename)
       
        else:
            sheet[name4].cell(row=lg_rw, column=15).value = sheet[name4].cell(row=lg_rw, column=11).value
            sheet[name4].cell(row=lg_rw, column=14).value = lgst
            sheet[name4].cell(row=lg_rw, column=10).value = None
            sheet[name4].cell(row=lg_rw, column=11).value = None
            sheet[name4].cell(row=lg_rw, column=19).value = lgst
            sheet[name4].cell(row=lg_rw, column=20).value = sheet[name4].cell(row=lg_rw, column=22).value
            sheet[name4].cell(row=lg_rw, column=21).value = None
            sheet[name4].cell(row=lg_rw, column=22).value = None
            
            tt_deduct -= lgst
            sheet.save(filename)
        
    c_p = 0
    
    for p in p_l:
        net_p = sheet[name4].cell(row=46, column=p) 
        net_p.value = 0
        cal_p = sheet[name4].cell(row=47, column=p)
        design_p = sheet[name4].cell(row=48, column=p)   
        for fnl in range(3, 45):    
            if sheet[name4].cell(row=fnl, column=p).value != None:    
                net_p.value += sheet[name4].cell(row=fnl, column=p).value
        net_p.value /= round(sheet[name].cell(row=2,column=28).value,1)
        net_p.value = round(net_p.value,1)
        sheet[name4].cell(row=46, column=p_2[c_p]).value = net_p.value
        cal_p.value = round(1.3*net_p.value,1)
        sheet[name4].cell(row=47, column=p_2[c_p]).value = cal_p.value
        design_p.value = round(1.2*cal_p.value,1)
        sheet[name4].cell(row=48, column=p_2[c_p]).value = design_p.value
        
        sheet[name4].cell(row=49,column=p).value = int(math.ceil(design_p.value))
        sheet[name4].cell(row=49, column=p_2[c_p]).value = sheet[name4].cell(row=49,column=p).value
        c_p+=1
    
    sheet.save(filename)
    return sheet


def make_sum(file_name, name, num, excel, st, gd, a1, a2, dflt=True):
    alp = ["0","A","B","C","D","E","F","G","H","I", "J","K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    num_l = [[3,7],[9,12],[14,19],[21,24],[26,27],[29,36],[38,44]]
    fill_pattern = styles.PatternFill(patternType='solid', fgColor='DDEBF7')
    count = -1
    for i in num:
        count+=1
        for j in range(st,gd):   
            
            form = "=SUM("+alp[j]+str(num_l[count][0])+":"+ alp[j]+str(num_l[count][1])+")"
            
            file_name[name].cell(row=i, column=j).value = form
            file_name[name].cell(row=i, column=j).font = styles.Font(bold=True)
            file_name[name].cell(row=i, column=j).fill = fill_pattern
            
        if dflt == True:    
            file_name[name].cell(row=i, column=gd).fill = fill_pattern
            for k in range(num_l[count][0]-1, num_l[count][1]+1):
                file_name[name].cell(row=k, column=gd).value="=SUM("+a1+str(k)+":"+a2+str(k)+")"
                file_name[name].cell(row=k, column=gd).font = styles.Font(bold=True)
        
    file_name.save(excel)
    return file_name

def grand_total(file_name, name, excel, num_l, cl, bg):
    
    cter = 0
    fill_pattern = styles.PatternFill(patternType='solid', fgColor='DDEBF7')
    alp = ["0","A","B","C","D","E","F","G","H","I", "J","K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    for i in num_l:
        form = "=SUM("+alp[i]+"2+"+alp[i]+"8+"+alp[i]+"13+"+alp[i]+"20+"+alp[i]+"25+"+alp[i]+"28+"+alp[i]+"37)"
        file_name[name].cell(row=45, column=i).value = form
        
        
    file_name.save(excel)
    return file_name



keywords = ["Air", "Body", "Bogie", "Brakes", "Cab", "Electrical", "Engine", "Grand Total"]
converted = make_tabs("test2.xlsx")
new_sheet = make_copy('accessrequired.xlsx', "test2.xlsx", 'Sheet1', converted, "Access Required", keywords)
new_sheet = make_copy('accessrequired.xlsx', "test2.xlsx", 'Sheet2', converted, "Infrastructure Required", keywords)
new_sheet = make_copy('accessrequired.xlsx', "test2.xlsx", 'Sheet3', converted, "Tool Required", keywords)
new_sheet = make_copy('accessrequired.xlsx', "test2.xlsx", 'Sheet4', converted, "Simple Workload", keywords)
new_sheet = make_copy('accessrequired.xlsx', "test2.xlsx", 'Sheet5', converted, "Data", keywords)
print("done making sheet")
final = data_transfer(new_sheet,"Sheet1", "Access Required", "Air", "test2.xlsx", [3,8])
final = data_transfer(final,"Sheet1", "Access Required", "Body", "test2.xlsx", [9,13])
final = data_transfer(final,"Sheet1", "Access Required", "Bogie", "test2.xlsx", [14,20])
final = data_transfer(final,"Sheet1", "Access Required", "Brakes", "test2.xlsx", [21,25])
final = data_transfer(final,"Sheet1", "Access Required", "Cab", "test2.xlsx", [26,28])
final = data_transfer(final,"Sheet1", "Access Required", "Electrical", "test2.xlsx", [29,37])
final = data_transfer(final,"Sheet1", "Access Required", "Engine", "test2.xlsx", [38,45])
final = inf_transfer(final,"Sheet1", "Infrastructure Required", "Air", "test2.xlsx", [3,8])
final = inf_transfer(final,"Sheet1", "Infrastructure Required", "Body", "test2.xlsx", [9,13])
final = inf_transfer(final,"Sheet1", "Infrastructure Required", "Bogie", "test2.xlsx", [14,20])
final = inf_transfer(final,"Sheet1", "Infrastructure Required", "Brakes", "test2.xlsx", [21,25])
final = inf_transfer(final,"Sheet1", "Infrastructure Required", "Cab", "test2.xlsx", [26,28])
final = inf_transfer(final,"Sheet1", "Infrastructure Required", "Electrical", "test2.xlsx", [29,37])
final = inf_transfer(final,"Sheet1", "Infrastructure Required", "Engine", "test2.xlsx", [38,45])
print("done inf data")
final = tool_transfer(final,"Sheet1", "Tool Required", "Air", "test2.xlsx", [3,8])
final = tool_transfer(final,"Sheet1", "Tool Required", "Body", "test2.xlsx", [9,13])
final = tool_transfer(final,"Sheet1", "Tool Required", "Bogie", "test2.xlsx", [14,20])
final = tool_transfer(final,"Sheet1", "Tool Required", "Brakes", "test2.xlsx", [21,25])
final = tool_transfer(final,"Sheet1", "Tool Required", "Cab", "test2.xlsx", [26,28])
final = tool_transfer(final,"Sheet1", "Tool Required", "Electrical", "test2.xlsx", [29,37])
final = tool_transfer(final,"Sheet1", "Tool Required", "Engine", "test2.xlsx", [38,45])
print("done tool data")
final = simp_wkl(final,"Sheet1", "Simple Workload", "Air", "test2.xlsx", [3,8], [10,2,2,2,2], ["/","/","/","/","/"])
final = simp_wkl(final,"Sheet1", "Simple Workload", "Body", "test2.xlsx", [9,13], [3.5,4,2,2,2], ["/","*","/","/"])
final = simp_wkl(final,"Sheet1", "Simple Workload", "Bogie", "test2.xlsx", [14,20], [3.8,4.5,8,3.75,1, 2], ["/","/","*","/","*","/"])
final = simp_wkl(final,"Sheet1", "Simple Workload", "Brakes", "test2.xlsx", [21,25], [12,4,1.2,2], ["/","*","/","/"])
final = simp_wkl(final,"Sheet1", "Simple Workload", "Cab", "test2.xlsx", [26,28], [7.75,2], ["/","/"])
final = simp_wkl(final,"Sheet1", "Simple Workload", "Electrical", "test2.xlsx", [29,37], [8,2,2,2,2,2,7,2], ["/","*","/","/","/","*","/","/"])
final = simp_wkl(final,"Sheet1", "Simple Workload", "Engine", "test2.xlsx", [38,45], [5.75,4,12,0.78,4,2.5,2.75], ["/","*","*","/","/","/","/"])
print("done simple data")
final = data(final, "Sheet1", "Infrastructure Required", "Simple Workload", "Data", "test2.xlsx", "Tool Required")
print("done dat data")
final = make_sum(final, "Access Required", [2,8,13,20,25,28,37], "test2.xlsx", 2, 11, "B", "J")
final = make_sum(final, "Infrastructure Required", [2,8,13,20,25,28,37], "test2.xlsx", 2, 5, "B", "D")
final = make_sum(final, "Infrastructure Required", [2,8,13,20,25,28,37], "test2.xlsx", 8, 11, "H", "J")
final = make_sum(final, "Infrastructure Required", [2,8,13,20,25,28,37], "test2.xlsx", 14, 17, "N", "P")
final = make_sum(final, "Tool Required", [2,8,13,20,25,28,37], "test2.xlsx", 2, 7, "B", "F")
final = make_sum(final, "Tool Required", [2,8,13,20,25,28,37], "test2.xlsx", 9, 14, "I", "M")
final = make_sum(final, "Tool Required", [2,8,13,20,25,28,37], "test2.xlsx", 17, 22, "Q", "U")
final = make_sum(final, "Simple Workload", [2,8,13,20,25,28,37], "test2.xlsx", 2, 5, "D", "D", False)
final = make_sum(final, "Simple Workload", [2,8,13,20,25,28,37], "test2.xlsx", 6, 8, "D", "D", False)
print("done make sun data")
final = grand_total(final, "Access Required", "test2.xlsx", [2,3,4,5,6,7,8,9,10,11],[11],[2])
final = grand_total(final, "Infrastructure Required", "test2.xlsx", [2,3,4,5,8,9,10,11,14,15,16,17],[5,11,17],[2,8,14])
final = grand_total(final, "Tool Required", "test2.xlsx", [2,3,4,5,6,7,9,10,11,12,13,14,17,18,19,20,21,22],[7,14,22],[2,9,17])
final = grand_total(final, "Simple Workload", "test2.xlsx", [2,3,4,6,7],[0],[0])
print("done grand data")


